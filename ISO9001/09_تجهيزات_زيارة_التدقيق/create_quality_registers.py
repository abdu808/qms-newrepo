# -*- coding: utf-8 -*-
"""
نظام سجلات الجودة المساند - الجمعية الخيرية
يُنشئ ملف Excel يحتوي على 5 أوراق عمل:
  1. لوحة المتابعة
  2. سجل المخاطر
  3. سجل الشكاوى والإجراءات التصحيحية
  4. سجل التدقيق الداخلي
  5. سجل تقييم الموردين

التشغيل:
  pip install openpyxl
  python create_quality_registers.py
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import datetime

# ألوان النظام
COLOR_HEADER     = "1F4E79"   # أزرق داكن
COLOR_SUBHEADER  = "2E75B6"   # أزرق متوسط
COLOR_ACCENT     = "BDD7EE"   # أزرق فاتح
COLOR_GREEN      = "E2EFDA"   # أخضر فاتح
COLOR_YELLOW     = "FFF2CC"   # أصفر فاتح
COLOR_RED        = "FCE4D6"   # أحمر فاتح
COLOR_WHITE      = "FFFFFF"
COLOR_GRAY       = "F2F2F2"
COLOR_DARK_TEXT  = "1F2D3D"

FONT_NAME = "Arial"

def thin_border():
    s = Side(style='thin', color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def header_style(ws, cell_ref, text, bg=COLOR_HEADER, fg=COLOR_WHITE, size=11, bold=True, merge_to=None):
    cell = ws[cell_ref]
    cell.value = text
    cell.font = Font(name=FONT_NAME, bold=bold, color=fg, size=size)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, reading_order=2)
    cell.border = thin_border()
    if merge_to:
        ws.merge_cells(f"{cell_ref}:{merge_to}")

def col_header(ws, row, cols_labels, bg=COLOR_SUBHEADER):
    for col_idx, label in enumerate(cols_labels, 1):
        cell = ws.cell(row=row, column=col_idx, value=label)
        cell.font = Font(name=FONT_NAME, bold=True, color=COLOR_WHITE, size=10)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, reading_order=2)
        cell.border = thin_border()

def data_row(ws, row, values, bg=COLOR_WHITE):
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = Font(name=FONT_NAME, size=10, color=COLOR_DARK_TEXT)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True, reading_order=2)
        cell.border = thin_border()

def set_rtl(ws):
    ws.sheet_view.rightToLeft = True

def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def freeze(ws, cell):
    ws.freeze_panes = cell


# ─────────────────────────────────────────────
# 1. لوحة المتابعة
# ─────────────────────────────────────────────
def create_dashboard(wb):
    ws = wb.active
    ws.title = "📊 لوحة المتابعة"
    set_rtl(ws)
    ws.row_dimensions[1].height = 50
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 40

    # عنوان رئيسي
    header_style(ws, "A1", "نظام سجلات الجودة المساند — الجمعية الخيرية",
                 bg=COLOR_HEADER, size=16, merge_to="H1")
    header_style(ws, "A2", f"آخر تحديث: {datetime.date.today().strftime('%Y/%m/%d')}",
                 bg=COLOR_SUBHEADER, size=10, merge_to="H2")

    # بطاقات الملخص
    cards = [
        ("A4", "D4", "⚠️ سجل المخاطر", COLOR_ACCENT),
        ("E4", "H4", "📋 سجل الشكاوى", COLOR_GREEN.replace("E2","C6")),
    ]
    header_style(ws, "A4", "⚠️ المخاطر النشطة", bg="2E75B6", size=11, merge_to="D4")
    header_style(ws, "E4", "📋 الشكاوى المفتوحة", bg="375623", size=11, merge_to="H4")

    # جدول الملخص
    ws.row_dimensions[6].height = 30
    header_style(ws, "A6", "السجل", bg=COLOR_SUBHEADER, size=10, merge_to="C6")
    header_style(ws, "D6", "إجمالي السجلات", bg=COLOR_SUBHEADER, size=10, merge_to="E6")
    header_style(ws, "F6", "مفتوح / قيد المعالجة", bg=COLOR_SUBHEADER, size=10, merge_to="G6")
    header_style(ws, "H6", "مغلق / منتهي", bg=COLOR_SUBHEADER, size=10)

    summary_rows = [
        ("سجل المخاطر", "راجع الورقة الثانية"),
        ("سجل الشكاوى والإجراءات التصحيحية", "راجع الورقة الثالثة"),
        ("سجل التدقيق الداخلي", "راجع الورقة الرابعة"),
        ("سجل تقييم الموردين", "راجع الورقة الخامسة"),
    ]
    for i, (name, note) in enumerate(summary_rows):
        r = 7 + i
        bg = COLOR_GRAY if i % 2 == 0 else COLOR_WHITE
        data_row(ws, r, [name, "", "", "—", "", "—", "", "—"], bg=bg)

    # تعليمات الاستخدام
    ws.row_dimensions[13].height = 25
    header_style(ws, "A13", "📌 تعليمات الاستخدام", bg=COLOR_HEADER, size=11, merge_to="H13")
    instructions = [
        "1. استخدم الأوراق المنفصلة لكل سجل.",
        "2. أدخل البيانات من اليمين إلى اليسار.",
        "3. حدّث حقل 'الحالة' فور أي تغيير.",
        "4. اعرض هذا الملف للمدقق دليلاً على التطبيق.",
        "5. احفظ نسخة احتياطية شهرياً.",
    ]
    for i, text in enumerate(instructions):
        r = 14 + i
        cell = ws.cell(row=r, column=1, value=text)
        cell.font = Font(name=FONT_NAME, size=10)
        cell.alignment = Alignment(horizontal="right", reading_order=2)
        ws.merge_cells(f"A{r}:H{r}")

    set_col_widths(ws, {"A":30,"B":5,"C":5,"D":12,"E":5,"F":18,"G":5,"H":12})


# ─────────────────────────────────────────────
# 2. سجل المخاطر
# ─────────────────────────────────────────────
def create_risks(wb):
    ws = wb.create_sheet("⚠️ سجل المخاطر")
    set_rtl(ws)
    ws.row_dimensions[1].height = 45
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 40

    header_style(ws, "A1", "سجل المخاطر والفرص — الجمعية الخيرية",
                 bg=COLOR_HEADER, size=14, merge_to="L1")
    header_style(ws, "A2", "المرجع: QP-003 | يُراجع: ربع سنوياً",
                 bg=COLOR_SUBHEADER, size=9, merge_to="L2")

    headers = [
        "رقم المخاطرة", "العملية المتأثرة", "وصف المخاطرة",
        "نوع\n(مخاطرة/فرصة)", "الاحتمال\n(1-3)", "التأثير\n(1-3)",
        "مستوى الخطورة\n(احتمال×تأثير)", "الإجراء الوقائي / الاستجابة",
        "المسؤول", "تاريخ المراجعة", "الحالة", "ملاحظات"
    ]
    col_header(ws, 3, headers)

    sample_data = [
        ["R-001", "خدمة المستفيدين", "تسرّب بيانات المستفيدين", "مخاطرة", 2, 3,
         "=E4*F4", "صلاحيات وصول محدودة + سياسة حماية البيانات", "مسؤول IT", "2026/01/01", "نشط", ""],
        ["R-002", "التبرعات", "نقص في التبرعات", "مخاطرة", 2, 3,
         "=E5*F5", "تنويع مصادر التمويل وبناء احتياطي", "المدير التنفيذي", "2026/01/01", "نشط", ""],
        ["R-003", "الموارد البشرية", "دوران عالٍ للموظفين", "مخاطرة", 2, 2,
         "=E6*F6", "تحسين بيئة العمل والتوثيق المعرفي", "مدير HR", "2026/01/01", "نشط", ""],
        ["R-004", "المشتريات", "موردون غير موثوقين", "مخاطرة", 1, 2,
         "=E7*F7", "تقييم الموردين قبل التعاقد + قائمة معتمدة", "مسؤول المشتريات", "2026/01/01", "نشط", ""],
        ["R-005", "التقنية", "انقطاع الأنظمة الإلكترونية", "مخاطرة", 1, 3,
         "=E8*F8", "نسخ احتياطية يومية + خطة استمرارية", "مسؤول IT", "2026/01/01", "نشط", ""],
        ["R-006", "التطوع", "فرصة: توسيع قاعدة المتطوعين", "فرصة", 3, 3,
         "=E9*F9", "برنامج استقطاب متطوعين + حملات توعية", "مسؤول التطوع", "2026/01/01", "نشط", ""],
    ]

    for i, row_data in enumerate(sample_data):
        r = 4 + i
        bg = COLOR_YELLOW if row_data[3] == "مخاطرة" and int(row_data[4])*int(row_data[5]) >= 6 else (COLOR_GREEN if row_data[3] == "فرصة" else COLOR_WHITE)
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = Font(name=FONT_NAME, size=10)
            cell.fill = PatternFill("solid", fgColor=bg.replace("#",""))
            cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True, reading_order=2)
            cell.border = thin_border()

    # صفوف فارغة
    for i in range(10, 25):
        bg = COLOR_GRAY if i % 2 == 0 else COLOR_WHITE
        for c in range(1, 13):
            cell = ws.cell(row=i, column=c, value="")
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="right", reading_order=2)

    # مفتاح الألوان
    ws.row_dimensions[26].height = 25
    header_style(ws, "A26", "مفتاح مستوى الخطورة:", bg=COLOR_HEADER, size=10, merge_to="C26")
    color_key = [("D26","E26","منخفض (1-2)", COLOR_GREEN),
                 ("F26","G26","متوسط (3-4)", COLOR_YELLOW),
                 ("H26","I26","عالٍ (6-9)", COLOR_RED)]
    for s, e, label, color in color_key:
        header_style(ws, s, label, bg=color.replace("#",""), fg=COLOR_DARK_TEXT, size=10, merge_to=e)

    set_col_widths(ws, {"A":12,"B":22,"C":35,"D":14,"E":12,"F":12,"G":16,"H":38,"I":18,"J":16,"K":12,"L":20})
    freeze(ws, "A4")


# ─────────────────────────────────────────────
# 3. سجل الشكاوى والإجراءات التصحيحية
# ─────────────────────────────────────────────
def create_complaints(wb):
    ws = wb.create_sheet("📋 الشكاوى والإجراءات التصحيحية")
    set_rtl(ws)
    ws.row_dimensions[1].height = 45
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 40

    header_style(ws, "A1", "سجل الشكاوى والإجراءات التصحيحية — الجمعية الخيرية",
                 bg=COLOR_HEADER, size=14, merge_to="M1")
    header_style(ws, "A2", "المرجع: CA-001 / CA-002 / QS-004 | الهدف: إغلاق الشكوى خلال 7 أيام عمل",
                 bg=COLOR_SUBHEADER, size=9, merge_to="M2")

    headers = [
        "رقم الحالة", "تاريخ الاستلام", "مصدر الشكوى\n(مستفيد/متطوع/متبرع/آخر)",
        "وصف الشكوى / عدم المطابقة", "العملية المتأثرة",
        "التصحيح الفوري\n(ماذا عملنا الآن)", "السبب الجذري",
        "الإجراء التصحيحي\n(لمنع التكرار)", "المسؤول",
        "تاريخ الإغلاق المستهدف", "تاريخ الإغلاق الفعلي",
        "الحالة", "تقييم الفاعلية"
    ]
    col_header(ws, 3, headers)

    sample_data = [
        ["C-001", "2026/01/15", "مستفيد", "تأخر في معالجة الطلب أكثر من أسبوع",
         "خدمة المستفيدين", "الاعتذار وتسريع المعالجة",
         "غياب موظف مسؤول دون تفويض بديل",
         "إنشاء آلية تفويض رسمية عند الغياب",
         "مدير الخدمات", "2026/01/22", "2026/01/20", "مغلق", "فعّال"],
        ["C-002", "2026/02/05", "متبرع", "عدم استلام إيصال التبرع",
         "إدارة التبرعات", "إرسال الإيصال فوراً",
         "خلل في إعدادات البريد الإلكتروني التلقائي",
         "مراجعة إعدادات النظام + اختبار شهري",
         "مسؤول التقنية", "2026/02/12", "", "مفتوح", ""],
    ]

    for i, row_data in enumerate(sample_data):
        r = 4 + i
        bg = COLOR_GREEN if row_data[11] == "مغلق" else COLOR_YELLOW
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = Font(name=FONT_NAME, size=10)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True, reading_order=2)
            cell.border = thin_border()

    for i in range(6, 30):
        bg = COLOR_GRAY if i % 2 == 0 else COLOR_WHITE
        for c in range(1, 14):
            cell = ws.cell(row=i, column=c, value="")
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="right", reading_order=2)

    set_col_widths(ws, {"A":10,"B":14,"C":20,"D":35,"E":22,"F":30,"G":28,"H":35,"I":16,"J":18,"K":18,"L":12,"M":18})
    freeze(ws, "A4")


# ─────────────────────────────────────────────
# 4. سجل التدقيق الداخلي
# ─────────────────────────────────────────────
def create_audit(wb):
    ws = wb.create_sheet("🔍 سجل التدقيق الداخلي")
    set_rtl(ws)
    ws.row_dimensions[1].height = 45
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 40

    header_style(ws, "A1", "سجل التدقيق الداخلي — الجمعية الخيرية",
                 bg=COLOR_HEADER, size=14, merge_to="K1")
    header_style(ws, "A2", "المرجع: IA-001 / IA-002 | يُنفَّذ مرة على الأقل كل 6 أشهر",
                 bg=COLOR_SUBHEADER, size=9, merge_to="K2")

    headers = [
        "رقم التدقيق", "تاريخ التدقيق", "الإدارة / العملية المدقّقة",
        "المدقق الداخلي", "نوع الملاحظة\n(كبرى/صغرى/فرصة)",
        "وصف الملاحظة", "بند ISO المرجعي",
        "الإجراء المطلوب", "المسؤول",
        "الموعد النهائي", "الحالة"
    ]
    col_header(ws, 3, headers)

    sample_data = [
        ["IA-2026-001", "2026/03/10", "خدمة المستفيدين",
         "مدير الجودة", "صغرى",
         "لا يوجد توثيق لنتائج قياس رضا المستفيدين",
         "ISO 9.1.2", "إنشاء نموذج استبانة رضا وتوثيق النتائج",
         "مدير الخدمات", "2026/04/10", "قيد التنفيذ"],
        ["IA-2026-002", "2026/03/10", "الموارد البشرية",
         "مدير الجودة", "فرصة تحسين",
         "يمكن رقمنة ملفات التدريب بدلاً من الورق",
         "ISO 7.2", "الانتقال لحفظ سجلات التدريب في رافد",
         "مدير HR", "2026/05/01", "مفتوح"],
        ["IA-2026-003", "2026/03/11", "المشتريات",
         "مدير الجودة", "صغرى",
         "ثلاثة موردين لم يُقيَّموا خلال آخر 12 شهر",
         "ISO 8.4", "إعادة تقييم الموردين المتأخرين",
         "مسؤول المشتريات", "2026/03/31", "مغلق"],
    ]

    for i, row_data in enumerate(sample_data):
        r = 4 + i
        status = row_data[10]
        bg = COLOR_GREEN if status == "مغلق" else (COLOR_YELLOW if status == "قيد التنفيذ" else COLOR_RED)
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = Font(name=FONT_NAME, size=10)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True, reading_order=2)
            cell.border = thin_border()

    for i in range(7, 30):
        bg = COLOR_GRAY if i % 2 == 0 else COLOR_WHITE
        for c in range(1, 12):
            cell = ws.cell(row=i, column=c, value="")
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="right", reading_order=2)

    set_col_widths(ws, {"A":14,"B":14,"C":25,"D":18,"E":18,"F":38,"G":16,"H":35,"I":16,"J":16,"K":14})
    freeze(ws, "A4")


# ─────────────────────────────────────────────
# 5. سجل تقييم الموردين
# ─────────────────────────────────────────────
def create_suppliers(wb):
    ws = wb.create_sheet("🏪 سجل تقييم الموردين")
    set_rtl(ws)
    ws.row_dimensions[1].height = 45
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 40

    header_style(ws, "A1", "سجل تقييم الموردين — الجمعية الخيرية",
                 bg=COLOR_HEADER, size=14, merge_to="L1")
    header_style(ws, "A2", "المرجع: REG-001 / REG-002 / PUR-001 | يُقيَّم الموردون سنوياً",
                 bg=COLOR_SUBHEADER, size=9, merge_to="L2")

    # رأس تقييم
    header_style(ws, "E3", "معايير التقييم (من 10)", bg=COLOR_ACCENT, fg=COLOR_DARK_TEXT, size=10, merge_to="I3")

    headers = [
        "رقم المورد", "اسم المورد", "نوع الخدمة / المنتج",
        "تاريخ آخر تقييم",
        "الجودة\n(10)", "الالتزام\nبالمواعيد (10)",
        "السعر\n(10)", "التواصل\n(10)", "الالتزام\nبالعقد (10)",
        "المجموع\n(50)", "التصنيف\n(أ/ب/ج)",
        "قرار الاستمرار"
    ]
    col_header(ws, 4, headers)

    sample_data = [
        ["S-001", "شركة الطباعة والتصميم", "مطبوعات وهوية بصرية",
         "2026/01/01", 9, 8, 7, 9, 8, "=E5+F5+G5+H5+I5", "=IF(J5>=40,\"أ\",IF(J5>=30,\"ب\",\"ج\"))", "استمرار"],
        ["S-002", "مزوّد خدمات الكيترينغ", "تموين الفعاليات",
         "2026/01/01", 8, 9, 8, 7, 9, "=E6+F6+G6+H6+I6", "=IF(J6>=40,\"أ\",IF(J6>=30,\"ب\",\"ج\"))", "استمرار"],
        ["S-003", "شركة الصيانة والنظافة", "خدمات النظافة للمقر",
         "2026/01/01", 7, 7, 9, 8, 7, "=E7+F7+G7+H7+I7", "=IF(J7>=40,\"أ\",IF(J7>=30,\"ب\",\"ج\"))", "استمرار"],
        ["S-004", "مزوّد البرمجيات والتقنية", "دعم تقني ونظم معلومات",
         "2026/01/01", 6, 5, 7, 5, 6, "=E8+F8+G8+H8+I8", "=IF(J8>=40,\"أ\",IF(J8>=30,\"ب\",\"ج\"))", "مراجعة"],
    ]

    for i, row_data in enumerate(sample_data):
        r = 5 + i
        score_col_idx = 10
        bg = COLOR_WHITE
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = Font(name=FONT_NAME, size=10)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="center" if c >= 5 else "right",
                                       vertical="center", wrap_text=True, reading_order=2)
            cell.border = thin_border()

    for i in range(9, 25):
        bg = COLOR_GRAY if i % 2 == 0 else COLOR_WHITE
        for c in range(1, 13):
            cell = ws.cell(row=i, column=c, value="")
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="right", reading_order=2)

    # مفتاح التصنيف
    ws.row_dimensions[26].height = 25
    header_style(ws, "A26", "مفتاح التصنيف:", bg=COLOR_HEADER, size=10, merge_to="B26")
    keys = [
        ("C26","D26","أ: 40-50 (ممتاز)", COLOR_GREEN),
        ("E26","F26","ب: 30-39 (جيد)", COLOR_YELLOW),
        ("G26","H26","ج: أقل من 30 (يحتاج مراجعة)", COLOR_RED),
    ]
    for s, e, label, color in keys:
        header_style(ws, s, label, bg=color, fg=COLOR_DARK_TEXT, size=10, merge_to=e)

    set_col_widths(ws, {"A":10,"B":28,"C":28,"D":16,"E":10,"F":14,"G":10,"H":12,"I":14,"J":12,"K":12,"L":18})
    freeze(ws, "A5")


# ─────────────────────────────────────────────
# التشغيل الرئيسي
# ─────────────────────────────────────────────
if __name__ == "__main__":
    try:
        from openpyxl import Workbook
    except ImportError:
        print("جارٍ تثبيت openpyxl...")
        import subprocess, sys
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        from openpyxl import Workbook

    wb = Workbook()

    create_dashboard(wb)
    create_risks(wb)
    create_complaints(wb)
    create_audit(wb)
    create_suppliers(wb)

    output_path = "نظام_سجلات_الجودة_المساند.xlsx"
    wb.save(output_path)
    print(f"\n✅ تم إنشاء الملف بنجاح: {output_path}")
    print("الأوراق المنشأة:")
    for sheet in wb.sheetnames:
        print(f"  • {sheet}")
